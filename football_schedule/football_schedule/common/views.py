from datetime import timedelta

import matplotlib
matplotlib.use("Agg")

from django.views.generic import TemplateView

import os
import zipfile
import tempfile
import logging
from dataclasses import dataclass
import matplotlib.pyplot as plt
import re
import math
import pandas as pd
from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from io import BytesIO


from football_schedule.schedules.models import Week, DisplayScheduleData

matplotlib.use("Agg")

# 🔥 GLOBAL matplotlib defaults (add THIS here)
plt.rcParams["savefig.format"] = "png"
plt.rcParams["savefig.dpi"] = 150


# Create your views here.
class HomeView(TemplateView):
    template_name = 'home.html'


def download_schedule_excel(request):
    # Fetch the schedule data for the logged-in user
    weeks = Week.objects.filter(author=request.user).order_by('id')

    # Prepare data for DataFrame
    data = []
    schedule_date = DisplayScheduleData.objects.last()
    data.append([schedule_date.club])
    data.append([schedule_date.team_generation])
    data.append([f"треньор {schedule_date.coach}"])
    data.append([f"Месец {schedule_date.month}"])

    for week in weeks:
        week_dates = [
            f"{week.start_date.strftime('%d.%m')}-{(week.start_date + timedelta(days=6)).strftime('%d.%m')}"
        ]
        # First row: Days of the week
        days_of_week = [
            "П", "В", "С", "Ч", "П", "С", "Н"
        ]

        types = [
            week.monday_type,
            week.tuesday_type,
            week.wednesday_type,
            week.thursday_type,
            week.friday_type,
            week.saturday_type,
            week.sunday_type,
        ]

        # Second row: Times for each day
        times = [
            f"{week.monday_time.strftime('%H:%M')} часа" if week.monday_time else "",
            f"{week.tuesday_time.strftime('%H:%M')} часа" if week.tuesday_time else "",
            f"{week.wednesday_time.strftime('%H:%M')} часа" if week.wednesday_time else "",
            f"{week.thursday_time.strftime('%H:%M')} часа" if week.thursday_time else "",
            f"{week.friday_time.strftime('%H:%M')} часа" if week.friday_time else "",
            f"{week.saturday_time.strftime('%H:%M')} часа" if week.saturday_time else "",
            f"{week.sunday_time.strftime('%H:%M')} часа" if week.sunday_time else "",
        ]

        # Third row: Places for each day
        places = [
            week.monday_place if week.monday_place else "",
            week.tuesday_place if week.tuesday_place else "",
            week.wednesday_place if week.wednesday_place else "",
            week.thursday_place if week.thursday_place else "",
            week.friday_place if week.friday_place else "",
            week.saturday_place if week.saturday_place else "",
            week.sunday_place if week.sunday_place else "",
        ]

        # Add the rows to data
        data.append(week_dates)
        data.append(days_of_week)
        data.append(types)
        data.append(times)
        data.append(places)

    # Create DataFrame
    df = pd.DataFrame(data)

    # Create an in-memory buffer to store the Excel file
    output = BytesIO()

    # Save DataFrame to Excel in memory
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Schedule", index=False, header=False)  # Do not include index or header

    # Load the workbook from the in-memory buffer
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active  # Get the active sheet (Schedule)

    merged_rows = [1,2,3,4,5,10,15,20,25]
    date_fill = PatternFill(start_color="0066ff",end_color="0066ff", fill_type="solid")
    day_fill = PatternFill(start_color="80b3ff",end_color="80b3ff", fill_type="solid")

    for row in merged_rows:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        if row > 4:
            ws.cell(row=row, column=1).fill = date_fill
            for c in range(1,8):
                ws.cell(row=row, column=c).fill = day_fill

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Apply border and center alignment to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Save the styled workbook back to the in-memory buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create HTTP response
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="schedule.xlsx"'

    return response





logger = logging.getLogger(__name__)

# ============================================================
# Process reports
# ============================================================

# CONSTRAINS

PLAYER_START_ROW = 5
PLAYER_END_ROW = 28

TRAINING_DATE_START = 0
TRAINING_DATE_END = 15

MATCH_DATE_START = 16
MATCH_DATE_END = 21

TRAINING_DATA_START = 1
TRAINING_DATA_END = 16

MATCH_DATA_START = 16

OUTPUT_FOLDER = "01_"

ABSENCE_COLOR = "#c21919"
TRAINING_PRESENCE_COLOR = "#43c01d"
MATCH_PRESENCE_COLOR = "#B8CCE4"


MONTH_MAP = {
    # January
    "януари": ("януари", "01"),
    "ян": ("януари", "01"),
    "яну": ("януари", "01"),
    "january": ("януари", "01"),
    "jan": ("януари", "01"),

    # February
    "февруари": ("февруари", "02"),
    "фев": ("февруари", "02"),
    "февр": ("февруари", "02"),
    "february": ("февруари", "02"),
    "feb": ("февруари", "02"),

    # March
    "март": ("март", "03"),
    "мар": ("март", "03"),
    "march": ("март", "03"),
    "mar": ("март", "03"),

    # April
    "април": ("април", "04"),
    "апр": ("април", "04"),
    "april": ("април", "04"),
    "apr": ("април", "04"),

    # May
    "май": ("май", "05"),
    "may": ("май", "05"),

    # June
    "юни": ("юни", "06"),
    "юн": ("юни", "06"),
    "june": ("юни", "06"),
    "jun": ("юни", "06"),

    # July
    "юли": ("юли", "07"),
    "юл": ("юли", "07"),
    "july": ("юли", "07"),
    "jul": ("юли", "07"),

    # August
    "август": ("август", "08"),
    "авг": ("август", "08"),
    "august": ("август", "08"),
    "aug": ("август", "08"),

    # September
    "септември": ("септември", "09"),
    "сеп": ("септември", "09"),
    "септ": ("септември", "09"),
    "september": ("септември", "09"),
    "sep": ("септември", "09"),
    "sept": ("септември", "09"),

    # October
    "октомври": ("октомври", "10"),
    "окт": ("октомври", "10"),
    "october": ("октомври", "10"),
    "oct": ("октомври", "10"),

    # November
    "ноември": ("ноември", "11"),
    "ное": ("ноември", "11"),
    "ноем": ("ноември", "11"),
    "november": ("ноември", "11"),
    "nov": ("ноември", "11"),

    # December
    "декември": ("декември", "12"),
    "дек": ("декември", "12"),
    "december": ("декември", "12"),
    "dec": ("декември", "12"),
}



# CLASSES

@dataclass
class ReportMetadata:
    club_name: str
    generation: str
    month: str
    month_num: str


@dataclass
class TrainingStatistics:
    training_dates: list
    absence_count: int
    presence_count: int
    trainings_count: int
    players_count: int

    @property
    def average_players(self):
        if self.trainings_count == 0:
            return 0

        return round(self.presence_count / self.trainings_count)


@dataclass
class MatchStatistics:
    date: object
    present: int

    @property
    def absent(self):
        return 0


@dataclass
class PlayerAttendance:
    name: str

    training_presence_dates: list
    training_absence_dates: list

    match_presence_dates: list
    match_absence_dates: list


def process_xlsm(input_path: str, output_dir: str):


    os.makedirs(output_dir, exist_ok=True)

    output_path = os.path.join(output_dir, OUTPUT_FOLDER)
    os.makedirs(output_path, exist_ok=True)

    absence_file = load_excel(input_path)


    metadata = extract_metadata(absence_file)
    training_dates, match_dates = extract_dates(absence_file)
    players = extract_players(absence_file)

    if not players:
        logger.warning("No players found in Excel file.")


    # Training report

    if training_dates and players:
        training_stats = calculate_training_statistics(
            absence_file,
            training_dates,
            len(players),
        )

        create_training_chart(
            training_stats,
            metadata,
            output_path,
        )
    else:
        logger.warning("Training report skipped: no training dates or players.")


    # Match report

    if match_dates and players:
        match_stats = calculate_match_statistics(
            absence_file,
            match_dates,
        )

        create_match_chart(
            match_stats,
            metadata,
            len(players),
            output_path,
        )
    else:
        logger.warning("Match report skipped: no match dates or players.")



    # Player reports

    if players:
        player_data = extract_player_attendance(
            absence_file,
            players,
            training_dates,
            match_dates,
            metadata.month_num,
        )

        for player in player_data:
            create_player_chart(
                player,
                metadata,
                output_path,
            )

    return output_dir


def load_excel(input_path: str) -> pd.DataFrame:

    return pd.read_excel(
        input_path,
        engine="openpyxl",
    )


def extract_metadata(df: pd.DataFrame) -> ReportMetadata:

    club_name = str(df.columns[3])
    generation = str(df.iloc[0, 3])
    raw_month = str(df.iloc[1, 3])

    month, month_num = parse_month(raw_month)

    return ReportMetadata(
        club_name=club_name,
        generation=generation,
        month=month,
        month_num=month_num,
    )


def parse_month(raw_month: str) -> tuple[str, str]:

    month = raw_month.strip().lower()

    parts = re.split(
        r"[\s\.,;:/\-]+",
        month,
    )

    for part in parts:
        if part in MONTH_MAP:
            return MONTH_MAP[part]

    raise ValueError(
        f"Unknown month value in Excel file: {raw_month!r}"
    )


def extract_dates(df: pd.DataFrame) -> tuple[list, list]:

    all_dates = df.iloc[4, 3:]

    training_dates = (
        all_dates.iloc[
            TRAINING_DATE_START:TRAINING_DATE_END
        ]
        .dropna()
        .tolist()
    )

    match_dates = (
        all_dates.iloc[
            MATCH_DATE_START:MATCH_DATE_END
        ]
        .dropna()
        .tolist()
    )

    return training_dates, match_dates


def extract_players(df: pd.DataFrame) -> list:

    players = (
        df.iloc[
            PLAYER_START_ROW:PLAYER_END_ROW,
            1,
        ]
        .dropna()
        .tolist()
    )

    return players


# ===== TRAINING DATA =====


def get_training_dataframe(df: pd.DataFrame) -> pd.DataFrame:

    training_df = df.iloc[
        PLAYER_START_ROW:PLAYER_END_ROW,
        1:19,
    ].copy()

    training_df.columns = df.iloc[4, 1:19]

    # Remove unwanted column
    training_df = training_df.drop(
        training_df.columns[1],
        axis=1,
    )

    return training_df


def calculate_training_statistics(
    df: pd.DataFrame,
    training_dates: list,
    players_count: int,
) -> TrainingStatistics:

    training_df = get_training_dataframe(df)

    # Training attendance block
    block = training_df.iloc[:, 1:16]

    presence_count = int(
        (block == 1).sum().sum()
    )

    absence_count = int(
        (block == 0).sum().sum()
    )

    trainings_count = len(training_dates)

    return TrainingStatistics(
        training_dates=training_dates,
        absence_count=absence_count,
        presence_count=presence_count,
        trainings_count=trainings_count,
        players_count=players_count,
    )


# ===== MATCH DATA =====

def get_match_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Extract the match section of the Excel file."""

    match_df = df.iloc[
        PLAYER_START_ROW:PLAYER_END_ROW,
        1:24,
    ].copy()

    match_df.columns = df.iloc[4, 1:24]

    # Keep player name + match columns
    match_df = match_df.iloc[
        :,
        [0] + list(range(18, 23)),
    ]

    return match_df


def calculate_match_statistics(
    df: pd.DataFrame,
    match_dates: list,
) -> list[MatchStatistics]:

    match_df = get_match_dataframe(df)

    # First column is player name.
    match_block = match_df.iloc[:, 1:]

    valid_columns = [
        column
        for column in match_block.columns
        if pd.notna(column)
    ]

    results = []

    for column in valid_columns:
        present = int(
            (match_block[column] == 1).sum()
        )

        results.append(
            MatchStatistics(
                date=column,
                present=present,
            )
        )

    return results



# ===== PLAYER =====

def get_players_dataframe(df: pd.DataFrame) -> pd.DataFrame:

    players_df = df.iloc[
        PLAYER_START_ROW:PLAYER_END_ROW,
        1:24,
    ].copy()

    players_df = players_df.drop(
        players_df.columns[1],
        axis=1,
    )

    players_df.reset_index(
        drop=True,
        inplace=True,
    )

    return players_df


def convert_attendance_values(values) -> list[int]:
    """
    Convert Excel attendance values to integers.

    Valid values:
        1 = present
        0 = absent

    Empty/invalid values become 0.
    """

    return (
        pd.to_numeric(
            values,
            errors="coerce",
        )
        .fillna(0)
        .astype(int)
        .tolist()
    )


def get_attendance_dates(
    events: list[int],
    dates: list,
    month_num: str,
) -> tuple[list[str], list[str]]:

    presence = []
    absence = []

    for event, date in zip(events, dates):

        if event == 1:
            presence.append(
                f"{date}.{month_num}"
            )

        elif event == 0:
            absence.append(
                f"{date}.{month_num}"
            )

    return presence, absence


def extract_player_attendance(
    df: pd.DataFrame,
    players: list,
    training_dates: list,
    match_dates: list,
    month_num: str,
) -> list[PlayerAttendance]:

    players_df = get_players_dataframe(df)

    result = []

    for index, player in enumerate(players):

        if index >= len(players_df):
            break

        player_row = players_df.iloc[index]

        player_name = str(
            player_row.iloc[0]
        ).strip()

        if not player_name or player_name.lower() == "nan":
            continue

        # Training attendance

        player_trainings = convert_attendance_values(
            player_row.iloc[
                TRAINING_DATA_START:
                TRAINING_DATA_END
            ]
        )

        training_presence_dates = []
        training_absence_dates = []

        if training_dates:
            (
                training_presence_dates,
                training_absence_dates,
            ) = get_attendance_dates(
                player_trainings,
                training_dates,
                month_num,
            )


        # Match attendance

        player_matches = convert_attendance_values(
            player_row.iloc[
                MATCH_DATA_START:
            ]
        )

        match_presence_dates = []
        match_absence_dates = []

        if match_dates:
            (
                match_presence_dates,
                match_absence_dates,
            ) = get_attendance_dates(
                player_matches,
                match_dates,
                month_num,
            )

        # Save result

        result.append(
            PlayerAttendance(
                name=player_name,
                training_presence_dates=training_presence_dates,
                training_absence_dates=training_absence_dates,
                match_presence_dates=match_presence_dates,
                match_absence_dates=match_absence_dates,
            )
        )

    return result


# MATPLOTLIB

def make_autopct(values, show_percent=True):
    """Create a Matplotlib autopct formatter."""

    total = sum(values)

    if total == 0:
        return lambda pct: "0"

    def formatter(pct):
        count = int(
            round(pct * total / 100.0)
        )

        if show_percent:
            return f"{pct:.0f}% ({count})"

        return str(count)

    return formatter


def save_figure(
    fig,
    output_path: str,
    dpi: int = 150,
):
    """Save and properly close a Matplotlib figure."""

    fig.savefig(
        output_path,
        dpi=dpi,
        bbox_inches="tight",
    )

    plt.close(fig)



# ===== CHART =====

def create_training_chart(
    stats: TrainingStatistics,
    metadata: ReportMetadata,
    output_path: str,
):

    labels = [
        "Отсъстващи",
        "Присъстващи",
    ]

    values = [
        stats.absence_count,
        stats.presence_count,
    ]

    fig, ax = plt.subplots(
        figsize=(6, 6)
    )

    wedges, _, autotexts = ax.pie(
        values,
        autopct=make_autopct(values),
        startangle=90,
        colors=[
            ABSENCE_COLOR,
            TRAINING_PRESENCE_COLOR,
        ],
    )

    for text in autotexts:
        text.set_fontsize(12)

    ax.set_title(
        f"{metadata.club_name}\n"
        f"{metadata.generation}\n"
        f"месец {metadata.month}"
    )

    ax.legend(
        wedges,
        labels,
        loc="upper right",
    )

    ax.text(
        0.5,
        0.05,
        f"Общо тренировки за месеца: "
        f"{stats.trainings_count}\n"
        f"Общо деца: {stats.players_count}\n"
        f"Средно деца на тренировка: "
        f"{stats.average_players}",
        ha="center",
        va="center",
        fontsize=12,
        color="black",
        bbox=dict(
            facecolor="white",
            alpha=0.7,
            edgecolor="none",
            pad=5,
        ),
        transform=ax.transAxes,
    )

    export_path = os.path.join(
        output_path,
        f"Информация_тренировки_"
        f"{metadata.month}.png",
    )

    save_figure(
        fig,
        export_path,
        dpi=150,
    )


# MATCH CHART

def create_match_chart(
    matches: list[MatchStatistics],
    metadata: ReportMetadata,
    players_count: int,
    output_path: str,
):

    if not matches:
        return

    labels = [
        "Отсъстващи",
        "Присъстващи",
    ]

    total_presence = sum(
        match.present
        for match in matches
    )

    total_absence = sum(
        players_count - match.present
        for match in matches
    )

    num_matches = len(matches)

    right_ncols = min(
        3,
        max(
            1,
            math.ceil(
                math.sqrt(num_matches)
            ),
        ),
    )

    right_nrows = math.ceil(
        num_matches / right_ncols
    )

    fig = plt.figure(
        figsize=(
            5 + 3 * right_ncols,
            3.5 * right_nrows,
        )
    )

    outer = fig.add_gridspec(
        nrows=max(1, right_nrows),
        ncols=2,
        width_ratios=[1.2, 2.0],
        wspace=0.25,
    )


    # LEFT: SUMMARY

    summary_ax = fig.add_subplot(
        outer[:, 0]
    )

    summary_values = [
        total_absence,
        total_presence,
    ]

    wedges, _, _ = summary_ax.pie(
        summary_values,
        autopct=make_autopct(
            summary_values,
            show_percent=True,
        ),
        startangle=90,
        colors=[
            ABSENCE_COLOR,
            MATCH_PRESENCE_COLOR,
        ],
        textprops={
            "fontsize": 11,
        },
    )

    summary_ax.legend(
        wedges,
        labels,
        loc="upper right",
    )

    summary_ax.set_title(
        f"Информация за всички мачове\n"
        f"Месец {metadata.month}",
        fontsize=14,
        pad=12,
    )

    summary_ax.text(
        0.5,
        0.05,
        f"Общо присъствали: "
        f"{total_presence}\n"
        f"Общо отсъствали: "
        f"{total_absence}",
        ha="center",
        va="center",
        fontsize=11,
        color="black",
        bbox=dict(
            facecolor="white",
            alpha=0.7,
            edgecolor="none",
            pad=5,
        ),
        transform=summary_ax.transAxes,
    )


    # RIGHT: INDIVIDUAL MATCHES

    right = outer[:, 1].subgridspec(
        nrows=right_nrows,
        ncols=right_ncols,
        hspace=0.35,
        wspace=0.30,
    )

    axes = [
        fig.add_subplot(right[i, j])
        for i in range(right_nrows)
        for j in range(right_ncols)
    ]

    for ax, match_number, match in zip(
        axes,
        range(1, num_matches + 1),
        matches,
    ):

        present = match.present
        absent = players_count - present

        values = [
            absent,
            present,
        ]

        ax.pie(
            values,
            autopct=make_autopct(
                values,
                show_percent=True,
            ),
            startangle=90,
            colors=[
                ABSENCE_COLOR,
                MATCH_PRESENCE_COLOR,
            ],
            textprops={
                "fontsize": 9,
            },
        )

        ax.set_title(
            f"Мач №{match_number} — "
            f"{match.date} "
            f"{metadata.month}",
            fontsize=10,
        )

        ax.text(
            0.5,
            0.05,
            f"Присъствали: {present}\n"
            f"Отсъствали: {absent}",
            ha="center",
            va="center",
            fontsize=6,
            color="black",
            bbox=dict(
                facecolor="white",
                alpha=0.65,
                edgecolor="none",
                pad=3,
            ),
            transform=ax.transAxes,
        )

    # Hide unused axes
    for ax in axes[num_matches:]:
        ax.axis("off")

    fig.tight_layout()

    export_path = os.path.join(
        output_path,
        f"Информация_мачове_"
        f"{metadata.month}.png",
    )

    save_figure(
        fig,
        export_path,
        dpi=150,
    )



# PLAYER CHART

def create_player_chart(
    player: PlayerAttendance,
    metadata: ReportMetadata,
    output_path: str,
):
    """Create the attendance chart for one player."""

    pies_to_draw = []

    if (
        player.training_presence_dates
        or player.training_absence_dates
    ):
        pies_to_draw.append(
            (
                "тренировки",
                player.training_presence_dates,
                player.training_absence_dates,
                TRAINING_PRESENCE_COLOR,
            )
        )

    if (
        player.match_presence_dates
        or player.match_absence_dates
    ):
        pies_to_draw.append(
            (
                "мачове",
                player.match_presence_dates,
                player.match_absence_dates,
                MATCH_PRESENCE_COLOR,
            )
        )

    if not pies_to_draw:
        return

    fig, axes = plt.subplots(
        1,
        len(pies_to_draw),
        figsize=(
            6 * len(pies_to_draw),
            6,
        ),
    )

    if len(pies_to_draw) == 1:
        axes = [axes]

    fig.suptitle(
        player.name,
        fontsize=18,
        fontweight="bold",
    )

    for ax, (
        event_type,
        presence,
        absence,
        presence_color,
    ) in zip(axes, pies_to_draw):

        values = [
            len(absence),
            len(presence),
        ]

        wedges, _, _ = ax.pie(
            values,
            autopct=make_autopct(
                values,
                show_percent=True,
            ),
            startangle=90,
            colors=[
                ABSENCE_COLOR,
                presence_color,
            ],
        )

        ax.set_title(
            f"{event_type} за месец "
            f"{metadata.month}"
        )

        ax.legend(
            wedges,
            [
                "Отсъствия",
                "Присъствия",
            ],
            loc="upper right",
        )

        summary = [
            f"Общо {event_type}: "
            f"{sum(values)}"
        ]

        if presence:
            summary.append(
                f"Присъствия: "
                f"{', '.join(map(str, presence))} "
                f"- общо {len(presence)}"
            )

        if absence:
            summary.append(
                f"Отсъствия: "
                f"{', '.join(map(str, absence))} "
                f"- общо {len(absence)}"
            )

        ax.text(
            0.5,
            0.1,
            "\n".join(summary),
            ha="center",
            va="top",
            fontsize=11,
            transform=ax.transAxes,
        )

    fig.tight_layout(
        rect=[0, 0, 1, 0.95]
    )

    export_path = os.path.join(
        output_path,
        f"{player.name}_"
        f"Информация_"
        f"{metadata.month}_месец.png",
    )

    save_figure(
        fig,
        export_path,
        dpi=300,
    )

